import logging
from tempfile import TemporaryFile

from datetime import datetime
from typing import Tuple, List
from urllib.parse import urljoin
from urllib.request import urlopen

from bs4 import BeautifulSoup
from cryptography.hazmat.primitives import hashes
from openpyxl import load_workbook, Workbook

from trust_stores_observatory.certificates_repository import RootCertificatesRepository
from trust_stores_observatory.store_fetcher.root_records_validator import RootRecordsValidator
from trust_stores_observatory.store_fetcher.scraped_root_record import ScrapedRootCertificateRecord
from trust_stores_observatory.store_fetcher.store_fetcher_interface import StoreFetcherInterface
from trust_stores_observatory.trust_store import TrustStore, PlatformEnum


class MicrosoftTrustStoreFetcher(StoreFetcherInterface):

    _INDEX_PAGE_URL = 'https://aka.ms/trustcertpartners'

    def fetch(self, certs_repo: RootCertificatesRepository, should_update_repo: bool=True) -> TrustStore:
        # Find the URL to the latest spreadsheet and download it
        spreadsheet_url = self._find_latest_root_certificates_url()
        with urlopen(spreadsheet_url) as response:
            spreadsheet_content = response.read()

        with TemporaryFile() as temp_file:
            temp_file.write(spreadsheet_content)
            workbook = load_workbook(temp_file)

        # Extract the data from the spreadsheet
        version, scraped_trusted_root_records, scraped_blocked_root_records = self._parse_spreadsheet(workbook)

        # Look for each parsed certificate in the supplied certs repo
        trusted_root_records = RootRecordsValidator.validate_with_repository(certs_repo, scraped_trusted_root_records)
        blocked_root_records = RootRecordsValidator.validate_with_repository(certs_repo, scraped_blocked_root_records)

        date_fetched = datetime.utcnow().date()
        return TrustStore(PlatformEnum.MICROSOFT_WINDOWS, version, spreadsheet_url, date_fetched, trusted_root_records,
                          blocked_root_records)

    @staticmethod
    def _parse_spreadsheet(
            workbook: Workbook
    ) -> Tuple[str, List[ScrapedRootCertificateRecord], List[ScrapedRootCertificateRecord]]:
        worksheet = workbook.active
        version = worksheet['A1'].value.split('As of')[1].strip()

        # Iterate over each row in the work sheet
        parsed_trusted_root_records = []
        parsed_blocked_root_records = []
        for row in worksheet.iter_rows(min_row=5, max_col=6, max_row=500):
            subject_name = row[1].value
            if subject_name is None:
                # Most likely indicates the end of the data
                continue

            is_cert_trusted = False
            status = row[4].value.strip()
            if 'Active' in status:
                # Some certs are disabled or have a notBefore constraint
                is_cert_trusted = True

            fingerprint_cell = row[3].value
            if fingerprint_cell is None:
                # One certificate actually does not have the fingerprint cell properly filled
                logging.error(f'No fingerprint for {subject_name}')
                continue

            fingerprint_hex = fingerprint_cell.replace(':', '').strip()
            fingerprint = bytes(bytearray.fromhex(fingerprint_hex))

            record = ScrapedRootCertificateRecord(subject_name, fingerprint, hashes.SHA256())
            if is_cert_trusted:
                parsed_trusted_root_records.append(record)
            else:
                parsed_blocked_root_records.append(record)

        return version, parsed_trusted_root_records, parsed_blocked_root_records

    @classmethod
    def _find_latest_root_certificates_url(cls) -> str:
        # Fetch and parse the index page
        with urlopen(cls._INDEX_PAGE_URL) as response:
            page_content = response.read()
        parsed_page = BeautifulSoup(page_content, 'html.parser')

        # Slow way to find the link
        next_page_url = None
        for p_tag in parsed_page.find_all('p'):
            if 'most recent list' in p_tag.text:
                next_page_url = p_tag.a['href']
                break
    
        if not next_page_url:
            raise ValueError(f'Could not find the next page URL at {cls._INDEX_PAGE_URL}')

        # Fetch and parse the next page which contains a link to a spreadsheet with the certificates
        with urlopen(next_page_url) as response:
            page_content = response.read()
        parsed_page = BeautifulSoup(page_content, 'html.parser')

        download_div = parsed_page.find('div', id='Downloads')
        spreadsheet_relative_url = download_div.a['href']
        spreadsheet_full_url = urljoin(next_page_url, spreadsheet_relative_url)
        return spreadsheet_full_url
